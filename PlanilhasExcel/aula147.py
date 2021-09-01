
import openpyxl

custos = openpyxl.load_workbook('Custos.xlsx')

nome_planilhas = custos.sheetnames

planilha_setembro = custos['Setembro']

# print(planilha_setembro['a3'].value)
#
# for campo in planilha_setembro['a']:
#     print(campo.value)
#
# for linha in planilha_setembro['a1:c2']:
#     for coluna in linha:
#         print(coluna.value)
#
# for linha in planilha_setembro:
#     for coluna in linha:
#         print(coluna.value)
#
# for linha in planilha_setembro:
#     i=0
#     while i <17:
#         if linha[i].value is not None:
#             print(linha[i].value, end=' ')
#         else:
#             break
#         i+=1
#     print()
#
# planilha_setembro['B3'].value = 30
# custos.save('nova_planilha.xlsx')

planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

from random import uniform

for linha in range(1, 15):
    numero_ordem = linha
    id = 1200 + linha
    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 1).value = numero_ordem
    planilha1.cell(linha, 2).value = id
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 15):
    planilha2.cell(linha, 1).value = f'Luiz {linha} {round(uniform(10, 100))}'
    planilha2.cell(linha, 2).value = f'Otávio {linha} {round(uniform(10, 100))}'
    planilha2.cell(linha, 3).value = f'Joãozinho {linha} {round(uniform(10, 100))}'

planilha.save('nova_planilha.xlsx')
